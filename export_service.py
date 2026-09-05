import io
import json
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ANALYSIS_VIDEO_COLUMNS, DIRECTION_SUMMARY_COLUMNS, FINAL_SCRIPT_COLUMNS
from common import clean_text

def format_sheet(
    worksheet,
):

    fill = PatternFill(
        "solid",
        fgColor="1F2937",
    )

    font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:

        cell.fill = fill
        cell.font = font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for index in range(
        1,
        worksheet.max_column + 1,
    ):

        worksheet.column_dimensions[
            get_column_letter(
                index
            )
        ].width = 23

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

def analysis_videos_to_df(
    result,
):

    rows = []

    for video in result.get(
        "videos",
        [],
    ):

        rows.append(
            {

                "视频编号":
                    video.get(
                        "video_index",
                        "",
                    ),

                "文件名":
                    video.get(
                        "filename",
                        "",
                    ),

                "一句话核心":
                    video.get(
                        "one_sentence_core",
                        "",
                    ),

                "该视频推理卖点":
                    "\n".join(
                        [
                            f"{i + 1}. {item}"
                            for i, item
                            in enumerate(
                                video.get(
                                    "inferred_selling_points",
                                    [],
                                )
                            )
                        ]
                    ),

                "爆款脚本路线":
                    video.get(
                        "script_route",
                        "",
                    ),

                "人群画像":
                    video.get(
                        "audience_profile",
                        "",
                    ),

                "年龄预估":
                    video.get(
                        "age_estimate",
                        "",
                    ),

                "前3秒Hook":
                    video.get(
                        "first_3s_hook",
                        "",
                    ),

                "画面与节奏":
                    video.get(
                        "visual_rhythm",
                        "",
                    ),

                "最值得吸收的3点":
                    "\n".join(
                        [
                            f"{i + 1}. {item}"
                            for i, item
                            in enumerate(
                                video.get(
                                    "top_absorb_points",
                                    [],
                                )
                            )
                        ]
                    ),

                "参考价值判断":
                    video.get(
                        "fit_reason",
                        "",
                    ),

                "推荐指数":
                    video.get(
                        "recommend_score",
                        "",
                    ),
            }
        )

    return pd.DataFrame(
        rows,
        columns=ANALYSIS_VIDEO_COLUMNS,
    )

def directions_summary_to_df(
    result,
):

    rows = []

    for direction in result.get(
        "directions",
        [],
    ):

        rows.append(
            {

                "方向":
                    direction.get(
                        "direction_name",
                        "",
                    ),

                "核心思路":
                    direction.get(
                        "core_idea",
                        "",
                    ),

                "目标人群":
                    direction.get(
                        "target_audience",
                        "",
                    ),

                "前3秒Hook":
                    direction.get(
                        "hook",
                        "",
                    ),

                "产品切入方式":
                    direction.get(
                        "product_entry",
                        "",
                    ),

                "推荐视角":
                    direction.get(
                        "recommended_perspective",
                        "",
                    ),

                "推荐小场景":
                    direction.get(
                        "recommended_scene",
                        "",
                    ),

                "可吸收点":
                    "\n".join(
                        [
                            f"- {item}"
                            for item
                            in direction.get(
                                "absorb_points",
                                [],
                            )
                        ]
                    ),

                "差异化点":
                    "\n".join(
                        [
                            f"- {item}"
                            for item
                            in direction.get(
                                "differentiation_points",
                                [],
                            )
                        ]
                    ),
            }
        )

    return pd.DataFrame(
        rows,
        columns=DIRECTION_SUMMARY_COLUMNS,
    )

def final_script_to_df(
    result,
):

    rows = []

    for shot in result.get(
        "storyboard",
        [],
    ):

        rows.append(
            {

                "分镜序号":
                    shot.get(
                        "sequence",
                        "",
                    ),

                "时间段":
                    shot.get(
                        "time_range",
                        "",
                    ),

                "机位/视角":
                    shot.get(
                        "shot",
                        "",
                    ),

                "画面描述(道具/动作)":
                    shot.get(
                        "visual",
                        "",
                    ),

                "手部动作":
                    shot.get(
                        "hand_action",
                        "",
                    ),

                "中文口播/字幕参考":
                    shot.get(
                        "copy_cn",
                        "",
                    ),

                "英文口播/字幕":
                    shot.get(
                        "copy_en",
                        "",
                    ),

                "音效/节奏提示":
                    shot.get(
                        "audio",
                        "",
                    ),

                "爆款吸收点":
                    shot.get(
                        "absorb_point",
                        "",
                    ),

                "差异化处理":
                    shot.get(
                        "difference_point",
                        "",
                    ),

                "设计目的(底层逻辑)":
                    shot.get(
                        "rationale",
                        "",
                    ),
            }
        )

    return pd.DataFrame(
        rows,
        columns=FINAL_SCRIPT_COLUMNS,
    )

def review_script_to_df(
    result,
):

    rows = []

    for shot in result.get(
        "optimized_script",
        [],
    ):

        rows.append(
            {

                "分镜序号":
                    shot.get(
                        "sequence",
                        "",
                    ),

                "时间段":
                    shot.get(
                        "time_range",
                        "",
                    ),

                "机位/视角":
                    shot.get(
                        "shot",
                        "",
                    ),

                "画面描述(道具/动作)":
                    shot.get(
                        "visual",
                        "",
                    ),

                "手部动作":
                    shot.get(
                        "hand_action",
                        "",
                    ),

                "中文口播/字幕参考":
                    shot.get(
                        "copy_cn",
                        "",
                    ),

                "英文口播/字幕":
                    shot.get(
                        "copy_en",
                        "",
                    ),

                "音效/节奏提示":
                    shot.get(
                        "audio",
                        "",
                    ),

                "爆款吸收点":
                    shot.get(
                        "absorb_point",
                        "",
                    ),

                "差异化处理":
                    shot.get(
                        "difference_point",
                        "",
                    ),

                "设计目的(底层逻辑)":
                    shot.get(
                        "rationale",
                        "",
                    ),
            }
        )

    return pd.DataFrame(
        rows,
        columns=FINAL_SCRIPT_COLUMNS,
    )

def build_analysis_export_excel(
    analysis_result,
    directions_result=None,
    final_script_result=None,
):

    output = io.BytesIO()

    summary = (
        analysis_result.get(
            "comparison_summary",
            {},
        )
    )

    summary_df = pd.DataFrame(
        [

            {
                "项目":
                    "一句话共同核心",

                "内容":
                    summary.get(
                        "one_sentence_core",
                        "",
                    ),
            },

            {
                "项目":
                    "共同爆款脚本路线",

                "内容":
                    summary.get(
                        "common_script_route",
                        "",
                    ),
            },

            {
                "项目":
                    "共同人群",

                "内容":
                    summary.get(
                        "common_audience",
                        "",
                    ),
            },

            {
                "项目":
                    "年龄预估",

                "内容":
                    summary.get(
                        "age_estimate",
                        "",
                    ),
            },

            {
                "项目":
                    "共同前3秒Hook",

                "内容":
                    summary.get(
                        "common_hook_pattern",
                        "",
                    ),
            },

            {
                "项目":
                    "共同画面与节奏",

                "内容":
                    summary.get(
                        "visual_rhythm",
                        "",
                    ),
            },

            {
                "项目":
                    "多视频关键差异",

                "内容":
                    summary.get(
                        "key_differences",
                        "",
                    ),
            },
        ]
    )

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        summary_df.to_excel(
            writer,
            index=False,
            sheet_name="爆款对比总结",
        )

        analysis_videos_to_df(
            analysis_result
        ).to_excel(
            writer,
            index=False,
            sheet_name="逐条视频拆解",
        )

        format_sheet(
            writer.book[
                "爆款对比总结"
            ]
        )

        format_sheet(
            writer.book[
                "逐条视频拆解"
            ]
        )

        if directions_result:

            directions_summary_to_df(
                directions_result
            ).to_excel(
                writer,
                index=False,
                sheet_name="3个方向概览",
            )

            format_sheet(
                writer.book[
                    "3个方向概览"
                ]
            )

        if final_script_result:

            final_script_to_df(
                final_script_result
            ).to_excel(
                writer,
                index=False,
                sheet_name="最终拍摄脚本",
            )

            format_sheet(
                writer.book[
                    "最终拍摄脚本"
                ]
            )

    output.seek(
        0
    )

    return output.getvalue()

def build_review_export_excel(
    review_result,
):

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        review_script_to_df(
            review_result
        ).to_excel(
            writer,
            index=False,
            sheet_name="优化版脚本",
        )

        format_sheet(
            writer.book[
                "优化版脚本"
            ]
        )

    output.seek(
        0
    )

    return output.getvalue()



def build_sop2_chatgpt_payload(deep_result, product_info, viral_name, own_name):
    return {
        "product": product_info,
        "viral_reference": {"filename": viral_name, "script_route": deep_result.get("viral_script_route", "")},
        "own_video": {"filename": own_name, "script_route": deep_result.get("own_script_route", "")},
        "core_gap": deep_result.get("core_gap", ""),
        "one_sentence_conclusion": deep_result.get("one_sentence_conclusion", ""),
        "strengths": deep_result.get("own_strengths", []),
        "weaknesses": deep_result.get("own_weaknesses", []),
        "reedit_value": deep_result.get("reedit_value", ""),
        "reedit_reason": deep_result.get("reedit_reason", ""),
        "keep_segments": deep_result.get("keep_segments", []),
        "delete_segments": deep_result.get("delete_segments", []),
        "move_segments": deep_result.get("move_segments", []),
        "reshoot_segments": deep_result.get("reshoot_segments", []),
        "editing_plan": deep_result.get("editing_plan", ""),
        "optimization_plan": deep_result.get("optimization_plan", ""),
        "comparison_dimensions": deep_result.get("comparison_dimensions", []),
    }


def build_sop2_export_excel(deep_result, chatgpt_payload):
    output = io.BytesIO()
    overview = pd.DataFrame([
        {"项目": "一句话结论", "内容": deep_result.get("one_sentence_conclusion", "")},
        {"项目": "爆款脚本路线", "内容": deep_result.get("viral_script_route", "")},
        {"项目": "我的脚本路线", "内容": deep_result.get("own_script_route", "")},
        {"项目": "核心差距", "内容": deep_result.get("core_gap", "")},
        {"项目": "重剪价值", "内容": deep_result.get("reedit_value", "")},
        {"项目": "重剪判断", "内容": deep_result.get("reedit_reason", "")},
        {"项目": "整体重剪计划", "内容": deep_result.get("editing_plan", "")},
        {"项目": "后续优化方向", "内容": deep_result.get("optimization_plan", "")},
    ])
    dims = pd.DataFrame(deep_result.get("comparison_dimensions", []))
    dims = dims.rename(columns={"dimension":"对比项","viral":"爆款视频","own":"我的作品","gap":"核心差距","suggestion":"建议"})
    sw_rows = []
    for x in deep_result.get("own_strengths", []): sw_rows.append({"类型":"我的优势","内容":x})
    for x in deep_result.get("own_weaknesses", []): sw_rows.append({"类型":"我的劣势","内容":x})
    strengths = pd.DataFrame(sw_rows, columns=["类型","内容"])
    edit_rows = []
    for x in deep_result.get("keep_segments", []): edit_rows.append({"动作":"保留","时间/镜头":x.get("time_range",""),"内容":x.get("content",""),"原因/目的":x.get("reason","")})
    for x in deep_result.get("delete_segments", []): edit_rows.append({"动作":"删除","时间/镜头":x.get("time_range",""),"内容":x.get("content",""),"原因/目的":x.get("reason","")})
    for x in deep_result.get("move_segments", []): edit_rows.append({"动作":"前移/调整","时间/镜头":f'{x.get("source_time","")} → {x.get("target_time","")}',"内容":x.get("content",""),"原因/目的":x.get("reason","")})
    for x in deep_result.get("reshoot_segments", []): edit_rows.append({"动作":"补拍","时间/镜头":x.get("shot",""),"内容":x.get("action",""),"原因/目的":x.get("purpose","")})
    edit_df = pd.DataFrame(edit_rows, columns=["动作","时间/镜头","内容","原因/目的"])
    json_df = pd.DataFrame([{"ChatGPT_JSON": json.dumps(chatgpt_payload, ensure_ascii=False, indent=2)}])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for df, name in [(overview,"对比总览"),(dims,"10维差距"),(strengths,"我的优势劣势"),(edit_df,"重剪与补拍计划"),(json_df,"ChatGPT_JSON")]:
            df.to_excel(writer, index=False, sheet_name=name)
            format_sheet(writer.book[name])
    output.seek(0)
    return output.getvalue()
